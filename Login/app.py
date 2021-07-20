import streamlit as st
import openpyxl as pxl
import pandas as pd
book = pxl.load_workbook('proj.xlsx')
sheet = book.active
maxrow = sheet.max_row+1
namestr = ""
emailid = ""
password = ""
namestr = st.text_input('Enter Your Name')
sheet.cell(row=maxrow, column=1).value = namestr
if(len(namestr) > 0):
    emailid = ""
    emailid = st.text_input('Enter Your Email Id')
    sheet.cell(row=maxrow-1, column=2).value = emailid
if(len(namestr) > 0 and len(emailid) > 0):
    password = ""
    password = st.text_input('Enter Your Password')
    sheet.cell(row=maxrow-2, column=3).value = password
    conf_password = st.text_input('Confirm Your Password')


book.save('proj.xlsx')
